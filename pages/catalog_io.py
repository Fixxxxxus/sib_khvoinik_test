"""Импорт/экспорт каталога в Excel (openpyxl)."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from django.db import transaction

from pages.models import CatalogCategory, Plant, PlantGalleryImage, PlantVariant


def export_catalog_workbook() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("Установите пакет openpyxl: pip install openpyxl") from e

    wb = Workbook()
    ws_c = wb.active
    ws_c.title = "categories"
    ws_c.append(
        ["slug", "label", "card_label", "description", "sort_order", "cover_path", "image_alt", "hub_links_json", "legacy_paths_json"]
    )
    for c in CatalogCategory.objects.order_by("sort_order", "label"):
        ws_c.append(
            [
                c.slug,
                c.label,
                c.card_label,
                c.description,
                c.sort_order,
                c.cover_path,
                c.image_alt,
                json.dumps(c.hub_links or [], ensure_ascii=False),
                json.dumps(c.legacy_paths or [], ensure_ascii=False),
            ]
        )

    ws_p = wb.create_sheet("plants")
    ws_p.append(
        [
            "slug",
            "name",
            "category_slug",
            "description",
            "cover_path",
            "image_alt",
            "height_hint",
            "frost",
            "light",
            "catalog_teaser_override",
            "is_new",
            "is_published",
            "also_in_category_slugs_json",
            "legacy_paths_json",
        ]
    )
    for p in Plant.objects.select_related("category").order_by("slug"):
        ws_p.append(
            [
                p.slug,
                p.name,
                p.category.slug if p.category_id else "",
                p.description,
                p.cover_path,
                p.image_alt,
                p.height_hint,
                p.frost,
                p.light,
                p.catalog_teaser_override,
                int(p.is_new),
                int(p.is_published),
                json.dumps(p.also_in_category_slugs or [], ensure_ascii=False),
                json.dumps(p.legacy_paths or [], ensure_ascii=False),
            ]
        )

    ws_v = wb.create_sheet("variants")
    ws_v.append(["plant_slug", "sort_order", "height", "container", "price", "in_stock"])
    for v in PlantVariant.objects.select_related("plant").order_by("plant__slug", "sort_order", "pk"):
        ws_v.append(
            [
                v.plant.slug if v.plant_id else "",
                v.sort_order,
                v.height,
                v.container,
                v.price,
                int(v.in_stock),
            ]
        )

    ws_g = wb.create_sheet("gallery")
    ws_g.append(["plant_slug", "sort_order", "image", "alt_text"])
    for g in PlantGalleryImage.objects.select_related("plant").order_by("plant__slug", "sort_order", "pk"):
        ws_g.append(
            [
                g.plant.slug if g.plant_id else "",
                g.sort_order,
                g.image.name if g.image else "",
                g.alt_text,
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_catalog_workbook(content: bytes, *, dry_run: bool = False) -> dict[str, int]:
    """
    Импорт из книги export_catalog_workbook().
    Галерея по имени файла из колонки image не восстанавливает бинарники — только строки-пути для справки;
    после импорта загрузите файлы через админку.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("Установите пакет openpyxl: pip install openpyxl") from e

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    stats = {"categories": 0, "plants": 0, "variants": 0}

    if "categories" in wb.sheetnames:
        ws = wb["categories"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not dry_run:
            with transaction.atomic():
                for row in rows:
                    if not row or not row[0]:
                        continue
                    slug, label = str(row[0]).strip(), str(row[1] or "").strip()
                    if not label:
                        continue
                    CatalogCategory.objects.update_or_create(
                        slug=slug,
                        defaults={
                            "label": label,
                            "card_label": str(row[2] or "").strip(),
                            "description": str(row[3] or "").strip(),
                            "sort_order": int(row[4] or 0),
                            "cover_path": str(row[5] or "").strip(),
                            "image_alt": str(row[6] or "").strip(),
                            "hub_links": json.loads(row[7] or "[]"),
                            "legacy_paths": json.loads(row[8] or "[]"),
                        },
                    )
                    stats["categories"] += 1
        else:
            stats["categories"] = sum(1 for r in rows if r and r[0])

    plant_slugs: dict[str, Plant] = {}
    if "plants" in wb.sheetnames:
        ws = wb["plants"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not dry_run:
            with transaction.atomic():
                for row in rows:
                    if not row or not row[0]:
                        continue
                    slug = str(row[0]).strip()
                    cat_slug = str(row[2] or "").strip()
                    cat = CatalogCategory.objects.filter(slug=cat_slug).first()
                    if not cat:
                        continue
                    p, _ = Plant.objects.update_or_create(
                        slug=slug,
                        defaults={
                            "name": str(row[1] or slug).strip(),
                            "category": cat,
                            "description": str(row[3] or "").strip() or "Черновик описания, заполните в админке.",
                            "cover_path": str(row[4] or "").strip(),
                            "image_alt": str(row[5] or "").strip(),
                            "height_hint": str(row[6] or "").strip(),
                            "frost": str(row[7] or "").strip(),
                            "light": str(row[8] or "").strip(),
                            "catalog_teaser_override": str(row[9] or "").strip(),
                            "is_new": bool(int(row[10] or 0)),
                            "is_published": bool(int(row[11] or 1)),
                            "also_in_category_slugs": json.loads(row[12] or "[]"),
                            "legacy_paths": json.loads(row[13] or "[]"),
                        },
                    )
                    plant_slugs[slug] = p
                    stats["plants"] += 1
        else:
            stats["plants"] = sum(1 for r in rows if r and r[0])

    if not dry_run and "variants" in wb.sheetnames:
        vrows = list(wb["variants"].iter_rows(min_row=2, values_only=True))
        v_slugs = {str(r[0]).strip() for r in vrows if r and r[0]}
        if v_slugs:
            PlantVariant.objects.filter(plant__slug__in=v_slugs).delete()
        with transaction.atomic():
            for row in vrows:
                if not row or not row[0]:
                    continue
                pslug = str(row[0]).strip()
                plant = plant_slugs.get(pslug) or Plant.objects.filter(slug=pslug).first()
                if not plant:
                    continue
                PlantVariant.objects.create(
                    plant=plant,
                    sort_order=int(row[1] or 0),
                    height=str(row[2] or "").strip(),
                    container=str(row[3] or "").strip(),
                    price=str(row[4] or "").strip(),
                    in_stock=bool(int(row[5] or 1)),
                )
                stats["variants"] += 1

    if dry_run:
        if "variants" in wb.sheetnames:
            stats["variants"] = sum(1 for r in wb["variants"].iter_rows(min_row=2, values_only=True) if r and r[0])

    return stats
